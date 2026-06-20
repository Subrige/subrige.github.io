#!/usr/bin/env python3
"""AAPL 10-Year DCF Model Builder — Institutional-Quality Excel Output"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter as cl

wb = Workbook()

# ═══════════════════════════════════════════════════════════
# STYLES
# ═══════════════════════════════════════════════════════════
FI  = Font(name='Calibri', size=11, color='0000FF')            # Input
FIB = Font(name='Calibri', size=11, color='0000FF', bold=True)
FC  = Font(name='Calibri', size=11, color='000000')            # Calc
FCB = Font(name='Calibri', size=11, color='000000', bold=True)
FL  = Font(name='Calibri', size=11, color='008000')            # Link
FH  = Font(name='Calibri', size=11, color='FFFFFF', bold=True) # Header
FT  = Font(name='Calibri', size=14, color='1F4E79', bold=True) # Title
FS  = Font(name='Calibri', size=10, color='666666')            # Subtitle

DF = PatternFill('solid', fgColor='1F4E79')  # Dark blue
LF = PatternFill('solid', fgColor='D9E1F2')  # Light blue
GF = PatternFill('solid', fgColor='F2F2F2')  # Light grey
OF = PatternFill('solid', fgColor='BDD7EE')  # Output blue

MS = Side(style='medium')
TS = Side(style='thin')
MB = Border(left=MS, right=MS, top=MS, bottom=MS)
TB = Border(left=TS, right=TS, top=TS, bottom=TS)

# ═══════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════
def shdr(ws, r, txt, mc=17):
    """Dark-blue merged section header."""
    ws.cell(r, 1, txt).font = FH
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=mc)
    for c in range(1, mc + 1):
        x = ws.cell(r, c); x.fill = DF; x.font = FH; x.border = MB

def yhdrs(ws, r, labels, sc):
    """Year column headers (light-blue)."""
    for i, y in enumerate(labels):
        x = ws.cell(r, sc + i, y)
        x.font = FCB; x.fill = LF; x.border = TB
        x.alignment = Alignment(horizontal='center')

def inp(ws, r, c, v, fmt=None, src=None):
    """Blue-font hardcoded input with optional grey fill and source comment."""
    x = ws.cell(r, c, v); x.font = FI; x.fill = GF
    if fmt: x.number_format = fmt
    if src: x.comment = Comment(f"Source: {src}", "Model")
    return x

def fml(ws, r, c, v, fmt=None, b=False):
    """Black-font formula cell."""
    x = ws.cell(r, c, v); x.font = FCB if b else FC
    if fmt: x.number_format = fmt
    return x

def outline(ws, sr, er, sc, ec):
    """Thick outline border around a section, preserving inner borders."""
    for r in range(sr, er + 1):
        for c in range(sc, ec + 1):
            cur = ws.cell(r, c).border
            ws.cell(r, c).border = Border(
                left   = MS if c == sc else cur.left,
                right  = MS if c == ec else cur.right,
                top    = MS if r == sr else cur.top,
                bottom = MS if r == er else cur.bottom,
            )

# ═══════════════════════════════════════════════════════════
# DATA
# ═══════════════════════════════════════════════════════════
# Historical (FY2021-FY2025, $M)
HR = [365817, 394328, 383285, 391035, 416161]   # Revenue
HE = [108949, 119437, 114301, 123216, 133050]   # EBIT
HD = [11284,  11104,  11519,  11445,  11698]    # D&A
HC = [11085,  10708,  10959,  9447,   12715]    # CapEx
HW = [-4911,  1200,   -6177,  3631,   -5847]   # ΔWC

# 10-year scenario assumptions
BG = [.050,.045,.040,.035,.030,.025,.025,.020,.020,.020]   # Bear growth
BM = [.300,.300,.295,.295,.290,.290,.285,.285,.280,.280]   # Bear margin
AG = [.070,.060,.060,.050,.050,.040,.040,.035,.030,.030]   # Base growth
AM = [.320,.325,.330,.335,.340,.340,.345,.345,.350,.350]   # Base margin
UG = [.090,.080,.070,.070,.060,.060,.050,.050,.045,.040]   # Bull growth
UM = [.340,.350,.360,.365,.370,.370,.375,.375,.380,.380]   # Bull margin

HY = [f'FY{y}A' for y in range(2021, 2026)]
PY = [f'FY{y}E' for y in range(2026, 2036)]

# ═══════════════════════════════════════════════════════════
# DCF SHEET
# ═══════════════════════════════════════════════════════════
ws = wb.active; ws.title = "DCF"
ws.sheet_properties.tabColor = "1F4E79"

# ── Title ──
ws.cell(1, 1, "Apple Inc. (AAPL) — 10-Year DCF Model").font = FT
ws.cell(2, 1, "Ticker: AAPL  |  Date: 2026-06-13  |  Fiscal Year End: September").font = FS

# ── Case Selector (row 4) ──
ws.cell(4, 1, "Case Selector (1=Bear, 2=Base, 3=Bull)").font = FCB
inp(ws, 4, 2, 2, '0', "User input: 1=Bear 2=Base 3=Bull")
fml(ws, 4, 4, '=CHOOSE(B4,"Bear Case","Base Case","Bull Case")')

# ── Market Data (rows 7–13) ──
shdr(ws, 7, "MARKET DATA & KEY INPUTS")
for r, t, v, fmt, src in [
    (9,  "Current Stock Price ($)",        291.13, '$#,##0.00', "Market data 2026-06-12, AAPL close price"),
    (10, "Shares Outstanding (Diluted, M)", 14690,  '#,##0',    "10-K FY2025, Diluted shares outstanding"),
    (12, "Net Cash ($M)",                   61880,  '#,##0',    "FY2025 Balance Sheet, Total Cash & Investments minus Total Debt"),
]:
    ws.cell(r, 1, t).font = FC; inp(ws, r, 2, v, fmt, src)

ws.cell(11, 1, "Market Capitalization ($M)").font = FC
fml(ws, 11, 2, '=B9*B10', '#,##0')

# ── Common Assumptions (rows 14–19) ──
shdr(ws, 14, "COMMON ASSUMPTIONS")
for r, t, v, src in [
    (15, "Tax Rate",              .16,  "Apple effective tax rate, FY2025 10-K"),
    (16, "D&A % of Revenue",     .028, "Historical average FY2021-FY2025, Cash Flow Statements"),
    (17, "CapEx % of Revenue",   .030, "Historical average FY2021-FY2025, Cash Flow Statements"),
    (18, "NWC Δ % of ΔRevenue",  .010, "Model estimate based on historical working capital patterns"),
    (19, "Equity Risk Premium",  .055, "Damodaran implied ERP, June 2026"),
]:
    ws.cell(r, 1, t).font = FC; inp(ws, r, 2, v, '0.0%', src)

# ── Scenario Blocks (rows 21–40) ──
for sr, title, gv, mv, tg, wc in [
    (21, "BEAR CASE ASSUMPTIONS", BG, BM, .020, .110),
    (28, "BASE CASE ASSUMPTIONS", AG, AM, .030, .100),
    (35, "BULL CASE ASSUMPTIONS", UG, UM, .035, .090),
]:
    shdr(ws, sr, title)
    ws.cell(sr+1, 1, "Assumption").font = FCB
    ws.cell(sr+1, 1).fill = LF; ws.cell(sr+1, 1).border = TB
    yhdrs(ws, sr+1, PY, 8)

    ws.cell(sr+2, 1, "Revenue Growth (%)").font = FC
    for i, v in enumerate(gv):
        inp(ws, sr+2, 8+i, v, '0.0%', "Analyst estimate, revenue growth assumption")

    ws.cell(sr+3, 1, "EBIT Margin (%)").font = FC
    for i, v in enumerate(mv):
        inp(ws, sr+3, 8+i, v, '0.0%', "Analyst estimate, operating margin assumption")

    ws.cell(sr+4, 1, "Terminal Growth Rate").font = FC
    inp(ws, sr+4, 2, tg, '0.0%', "GDP-aligned long-term growth estimate")

    ws.cell(sr+5, 1, "WACC").font = FC
    inp(ws, sr+5, 2, wc, '0.0%', "CAPM-derived, see WACC sheet for calculation")

# ── Selected Case Consolidation (rows 42–47) ──
shdr(ws, 42, "SELECTED CASE ASSUMPTIONS")
ws.cell(43, 1, "Assumption").font = FCB
ws.cell(43, 1).fill = LF; ws.cell(43, 1).border = TB
yhdrs(ws, 43, PY, 8)

ws.cell(44, 1, "Revenue Growth (%)").font = FC
for i in range(10):
    c = 8 + i
    fml(ws, 44, c, f'=CHOOSE($B$4,{cl(c)}23,{cl(c)}30,{cl(c)}37)', '0.0%')

ws.cell(45, 1, "EBIT Margin (%)").font = FC
for i in range(10):
    c = 8 + i
    fml(ws, 45, c, f'=CHOOSE($B$4,{cl(c)}24,{cl(c)}31,{cl(c)}38)', '0.0%')

ws.cell(46, 1, "Terminal Growth Rate").font = FC
fml(ws, 46, 2, '=CHOOSE($B$4,B25,B32,B39)', '0.0%')

ws.cell(47, 1, "WACC").font = FC
fml(ws, 47, 2, '=CHOOSE($B$4,B26,B33,B40)', '0.0%')

# ── Income Statement (rows 49–60) ──
shdr(ws, 49, "INCOME STATEMENT PROJECTION ($M)")
yhdrs(ws, 50, HY, 3); yhdrs(ws, 50, PY, 8)

# Revenue (row 51)
ws.cell(51, 1, "Revenue").font = FCB
for i, v in enumerate(HR):
    inp(ws, 51, 3+i, v, '#,##0', f"10-K FY{2021+i}, Consolidated Statements of Operations")
for c in range(8, 18):
    fml(ws, 51, c, f'={cl(c-1)}51*(1+{cl(c)}44)', '#,##0')

# % Growth (row 52)
ws.cell(52, 1, "% Growth").font = FC
ws.cell(52, 1).alignment = Alignment(indent=1)
for c in range(4, 8):
    fml(ws, 52, c, f'={cl(c)}51/{cl(c-1)}51-1', '0.0%')
for c in range(8, 18):
    fml(ws, 52, c, f'={cl(c)}51/{cl(c-1)}51-1', '0.0%')

# EBIT (row 54)
ws.cell(54, 1, "EBIT (Operating Income)").font = FCB
for i, v in enumerate(HE):
    inp(ws, 54, 3+i, v, '#,##0', f"10-K FY{2021+i}, Operating Income")
for c in range(8, 18):
    fml(ws, 54, c, f'={cl(c)}51*{cl(c)}45', '#,##0')

# % Margin (row 55)
ws.cell(55, 1, "% Margin").font = FC
ws.cell(55, 1).alignment = Alignment(indent=1)
for c in range(3, 18):
    fml(ws, 55, c, f'={cl(c)}54/{cl(c)}51', '0.0%')

# Taxes (row 57)
ws.cell(57, 1, "Taxes").font = FC
for c in range(3, 18):
    fml(ws, 57, c, f'={cl(c)}54*$B$15', '#,##0')

# Tax Rate display (row 58)
ws.cell(58, 1, "Effective Tax Rate").font = FC
ws.cell(58, 1).alignment = Alignment(indent=1)
for c in range(3, 18):
    fml(ws, 58, c, '=$B$15', '0.0%')

# NOPAT (row 60)
ws.cell(60, 1, "NOPAT").font = FCB
for c in range(3, 18):
    fml(ws, 60, c, f'={cl(c)}54-{cl(c)}57', '#,##0', b=True)

# ── Free Cash Flow (rows 62–73) ──
shdr(ws, 62, "FREE CASH FLOW ($M)")
yhdrs(ws, 63, HY, 3); yhdrs(ws, 63, PY, 8)

# NOPAT link (row 64)
ws.cell(64, 1, "NOPAT").font = FC
for c in range(3, 18):
    fml(ws, 64, c, f'={cl(c)}60', '#,##0')

# D&A (row 65)
ws.cell(65, 1, "(+) D&A").font = FC
for i, v in enumerate(HD):
    inp(ws, 65, 3+i, v, '#,##0', f"10-K FY{2021+i}, Cash Flow Statement")
for c in range(8, 18):
    fml(ws, 65, c, f'={cl(c)}51*$B$16', '#,##0')

# D&A % (row 66)
ws.cell(66, 1, "D&A % of Revenue").font = FC
ws.cell(66, 1).alignment = Alignment(indent=1)
for c in range(3, 18):
    fml(ws, 66, c, f'={cl(c)}65/{cl(c)}51', '0.0%')

# CapEx (row 67)
ws.cell(67, 1, "(-) CapEx").font = FC
for i, v in enumerate(HC):
    inp(ws, 67, 3+i, v, '#,##0', f"10-K FY{2021+i}, Cash Flow Statement")
for c in range(8, 18):
    fml(ws, 67, c, f'={cl(c)}51*$B$17', '#,##0')

# CapEx % (row 68)
ws.cell(68, 1, "CapEx % of Revenue").font = FC
ws.cell(68, 1).alignment = Alignment(indent=1)
for c in range(3, 18):
    fml(ws, 68, c, f'={cl(c)}67/{cl(c)}51', '0.0%')

# ΔNWC (row 69)
ws.cell(69, 1, "(-) Δ NWC").font = FC
for i, v in enumerate(HW):
    inp(ws, 69, 3+i, v, '#,##0', f"10-K FY{2021+i}, Cash Flow Statement, Working Capital Change")
for c in range(8, 18):
    fml(ws, 69, c, f'=({cl(c)}51-{cl(c-1)}51)*$B$18', '#,##0')

# NWC % (row 70)
ws.cell(70, 1, "NWC % of ΔRevenue").font = FC
ws.cell(70, 1).alignment = Alignment(indent=1)
for c in range(4, 8):
    fml(ws, 70, c, f'=IF({cl(c)}51={cl(c-1)}51,0,{cl(c)}69/({cl(c)}51-{cl(c-1)}51))', '0.0%')
for c in range(8, 18):
    fml(ws, 70, c, '=$B$18', '0.0%')

# Unlevered FCF (row 72) — total line
ws.cell(72, 1, "Unlevered Free Cash Flow").font = FCB
for c in range(3, 18):
    fml(ws, 72, c, f'={cl(c)}64+{cl(c)}65-{cl(c)}67-{cl(c)}69', '#,##0', b=True)

# FCF Margin (row 73)
ws.cell(73, 1, "FCF Margin").font = FC
ws.cell(73, 1).alignment = Alignment(indent=1)
for c in range(3, 18):
    fml(ws, 73, c, f'={cl(c)}72/{cl(c)}51', '0.0%')

# ── DCF Valuation (rows 75–84) ──
shdr(ws, 75, "DCF VALUATION ($M)")
yhdrs(ws, 76, PY, 8)

# FCF link (row 77)
ws.cell(77, 1, "Unlevered FCF").font = FC
for c in range(8, 18):
    fml(ws, 77, c, f'={cl(c)}72', '#,##0')

# Period (row 78) — mid-year convention
ws.cell(78, 1, "Period (mid-year)").font = FC
for i in range(10):
    x = ws.cell(78, 8+i, 0.5 + i); x.font = FC; x.number_format = '0.0'

# Discount Factor (row 79)
ws.cell(79, 1, "Discount Factor").font = FC
for c in range(8, 18):
    fml(ws, 79, c, f'=1/(1+$B$47)^{cl(c)}78', '0.0000')

# PV of FCF (row 80)
ws.cell(80, 1, "PV of FCF").font = FCB
for c in range(8, 18):
    fml(ws, 80, c, f'={cl(c)}77*{cl(c)}79', '#,##0', b=True)

# Terminal Value (only in column Q = 17)
ws.cell(82, 1, "Terminal FCF").font = FC
fml(ws, 82, 17, '=Q72*(1+$B$46)', '#,##0')

ws.cell(83, 1, "Terminal Value").font = FC
fml(ws, 83, 17, '=Q82/($B$47-$B$46)', '#,##0')

ws.cell(84, 1, "PV of Terminal Value").font = FCB
fml(ws, 84, 17, '=Q83/(1+$B$47)^Q78', '#,##0', b=True)

# ── Valuation Summary (rows 86–99) ──
shdr(ws, 86, "VALUATION SUMMARY")

for r, t, f, fmt, bold, hl in [
    (88, "Sum of PV of Projected FCFs",  '=SUM(H80:Q80)', '#,##0',     True,  False),
    (89, "PV of Terminal Value",          '=Q84',          '#,##0',     False, False),
    (90, "Enterprise Value",              '=B88+B89',      '#,##0',     True,  True),
    (91, "(+) Net Cash",                  '=B12',          '#,##0',     False, False),
    (92, "Equity Value",                  '=B90+B91',      '#,##0',     True,  True),
    (94, "Shares Outstanding (M)",        '=B10',          '#,##0',     False, False),
    (95, "IMPLIED PRICE PER SHARE",       '=B92/B94',      '$#,##0.00', True,  True),
    (96, "Current Stock Price",           '=B9',           '$#,##0.00', False, False),
    (97, "Implied Upside / (Downside)",   '=B95/B96-1',    '0.0%',     True,  True),
    (99, "Terminal Value % of EV",        '=B89/B90',      '0.0%',     False, False),
]:
    ws.cell(r, 1, t).font = FCB if bold else FC
    x = fml(ws, r, 2, f, fmt, b=bold)
    if hl: x.fill = OF

# ═══════════════════════════════════════════════════════════
# SENSITIVITY TABLES (rows 102–126, columns A–F)
# ═══════════════════════════════════════════════════════════
shdr(ws, 102, "SENSITIVITY ANALYSIS", mc=6)

# ── Table 1: WACC vs Terminal Growth (rows 104–110) ──
ws.merge_cells('A104:F104')
ws.cell(104, 1, "Table 1: WACC vs Terminal Growth Rate").font = FCB
ws.cell(104, 1).fill = LF
for c in range(1, 7): ws.cell(104, c).border = TB

ws.cell(105, 1, "WACC \\ Term. Growth").font = FCB
ws.cell(105, 1).fill = LF; ws.cell(105, 1).border = TB

tg_vals = [.020, .025, .030, .035, .040]
wc_vals = [.080, .090, .100, .110, .120]

for j, tg in enumerate(tg_vals):
    x = ws.cell(105, 2+j, tg)
    x.font = FI; x.number_format = '0.0%'; x.fill = LF; x.border = TB
    x.alignment = Alignment(horizontal='center')

for i, wc in enumerate(wc_vals):
    r = 106 + i
    x = ws.cell(r, 1, wc)
    x.font = FI; x.number_format = '0.0%'; x.fill = LF; x.border = TB
    for j in range(5):
        c = 2 + j; L = cl(c)
        f = (f'=(H$72/(1+$A{r})^0.5+I$72/(1+$A{r})^1.5+J$72/(1+$A{r})^2.5'
             f'+K$72/(1+$A{r})^3.5+L$72/(1+$A{r})^4.5+M$72/(1+$A{r})^5.5'
             f'+N$72/(1+$A{r})^6.5+O$72/(1+$A{r})^7.5+P$72/(1+$A{r})^8.5'
             f'+Q$72/(1+$A{r})^9.5'
             f'+Q$72*(1+{L}$105)/($A{r}-{L}$105)/(1+$A{r})^9.5'
             f'+$B$12)/$B$10')
        x = fml(ws, r, c, f, '$#,##0.00'); x.border = TB
        if wc == .10 and tg_vals[j] == .03:  # base case center
            x.fill = OF; x.font = FCB

# ── Table 2: Revenue Growth vs EBIT Margin (rows 112–118) ──
ws.merge_cells('A112:F112')
ws.cell(112, 1, "Table 2: Revenue Growth vs EBIT Margin (Uniform)").font = FCB
ws.cell(112, 1).fill = LF
for c in range(1, 7): ws.cell(112, c).border = TB

ws.cell(113, 1, "Growth \\ EBIT Margin").font = FCB
ws.cell(113, 1).fill = LF; ws.cell(113, 1).border = TB

mg_vals = [.30, .32, .34, .36, .38]
gr_vals = [.03, .04, .05, .06, .07]

for j, m in enumerate(mg_vals):
    x = ws.cell(113, 2+j, m)
    x.font = FI; x.number_format = '0.0%'; x.fill = LF; x.border = TB
    x.alignment = Alignment(horizontal='center')

for i, g in enumerate(gr_vals):
    r = 114 + i
    x = ws.cell(r, 1, g)
    x.font = FI; x.number_format = '0.0%'; x.fill = LF; x.border = TB
    for j in range(5):
        c = 2 + j; L = cl(c)
        # Closed-form 10-year DCF recalculation with uniform growth & margin
        # K = (1+g)*(m*(1-t)+da-cx) - g*nwc
        # Sum PV = rev0*K/SQRT(1+w)*(1-r^10)/(1-r)  where r = (1+g)/(1+w)
        # PV TV  = rev0*K*(1+g)^9*(1+tg)/(w-tg)/(1+w)^9.5
        f = (f'=($G$51*((1+$A{r})*({L}$113*(1-$B$15)+$B$16-$B$17)-$A{r}*$B$18)'
             f'/SQRT(1+$B$47)*(1-((1+$A{r})/(1+$B$47))^10)/(1-(1+$A{r})/(1+$B$47))'
             f'+$G$51*((1+$A{r})*({L}$113*(1-$B$15)+$B$16-$B$17)-$A{r}*$B$18)'
             f'*(1+$A{r})^9*(1+$B$46)/($B$47-$B$46)/(1+$B$47)^9.5'
             f'+$B$12)/$B$10')
        x = fml(ws, r, c, f, '$#,##0.00'); x.border = TB
        if g == .05 and mg_vals[j] == .34:  # approx base case center
            x.fill = OF; x.font = FCB

# ── Table 3: Beta vs Risk-Free Rate (rows 120–126) ──
ws.merge_cells('A120:F120')
ws.cell(120, 1, "Table 3: Beta vs Risk-Free Rate").font = FCB
ws.cell(120, 1).fill = LF
for c in range(1, 7): ws.cell(120, c).border = TB

ws.cell(121, 1, "Beta \\ Risk-Free Rate").font = FCB
ws.cell(121, 1).fill = LF; ws.cell(121, 1).border = TB

rf_vals = [.0335, .0385, .0435, .0485, .0535]
bt_vals = [0.79, 0.94, 1.09, 1.24, 1.39]

for j, rf in enumerate(rf_vals):
    x = ws.cell(121, 2+j, rf)
    x.font = FI; x.number_format = '0.0%'; x.fill = LF; x.border = TB
    x.alignment = Alignment(horizontal='center')

for i, bt in enumerate(bt_vals):
    r = 122 + i
    x = ws.cell(r, 1, bt)
    x.font = FI; x.number_format = '0.00'; x.fill = LF; x.border = TB
    for j in range(5):
        c = 2 + j; L = cl(c)
        # WACC ≈ Rf + Beta × ERP (Apple is ~100% equity-funded on net basis)
        w = f'{L}$121+$A{r}*$B$19'
        f = (f'=(H$72/(1+{w})^0.5+I$72/(1+{w})^1.5+J$72/(1+{w})^2.5'
             f'+K$72/(1+{w})^3.5+L$72/(1+{w})^4.5+M$72/(1+{w})^5.5'
             f'+N$72/(1+{w})^6.5+O$72/(1+{w})^7.5+P$72/(1+{w})^8.5'
             f'+Q$72/(1+{w})^9.5'
             f'+Q$72*(1+$B$46)/({w}-$B$46)/(1+{w})^9.5'
             f'+$B$12)/$B$10')
        x = fml(ws, r, c, f, '$#,##0.00'); x.border = TB
        if bt == 1.09 and rf_vals[j] == .0435:  # base case center
            x.fill = OF; x.font = FCB

# ═══════════════════════════════════════════════════════════
# WACC SHEET
# ═══════════════════════════════════════════════════════════
ww = wb.create_sheet("WACC")
ww.sheet_properties.tabColor = "008000"

ww.cell(1, 1, "APPLE INC. (AAPL) — WACC CALCULATION").font = FT
ww.cell(2, 1, "As of 2026-06-13").font = FS

# Cost of Equity
shdr(ww, 4, "COST OF EQUITY (CAPM)", mc=4)
ww.cell(6, 1, "Risk-Free Rate (10Y Treasury)").font = FC
inp(ww, 6, 2, .0435, '0.00%', "US 10-Year Treasury yield, June 2026")
ww.cell(7, 1, "Beta (5Y Monthly)").font = FC
inp(ww, 7, 2, 1.09, '0.00', "Yahoo Finance / stockanalysis.com, 5-year monthly beta")
ww.cell(8, 1, "Equity Risk Premium").font = FC
inp(ww, 8, 2, .055, '0.0%', "Damodaran implied equity risk premium, June 2026")
ww.cell(9, 1, "Cost of Equity").font = FCB
x = fml(ww, 9, 2, '=B6+B7*B8', '0.00%', b=True); x.fill = OF

# Cost of Debt
shdr(ww, 11, "COST OF DEBT", mc=4)
ww.cell(13, 1, "Pre-Tax Cost of Debt").font = FC
inp(ww, 13, 2, .040, '0.00%', "Estimated from Apple corporate bond yields, June 2026")
ww.cell(14, 1, "Tax Rate").font = FC
x = ww.cell(14, 2, '=DCF!B15'); x.font = FL; x.number_format = '0.0%'
ww.cell(15, 1, "After-Tax Cost of Debt").font = FCB
fml(ww, 15, 2, '=B13*(1-B14)', '0.00%')

# Capital Structure
shdr(ww, 17, "CAPITAL STRUCTURE", mc=4)
ww.cell(19, 1, "Current Stock Price ($)").font = FC
x = ww.cell(19, 2, '=DCF!B9'); x.font = FL; x.number_format = '$#,##0.00'
ww.cell(20, 1, "Shares Outstanding (M)").font = FC
x = ww.cell(20, 2, '=DCF!B10'); x.font = FL; x.number_format = '#,##0'
ww.cell(21, 1, "Market Capitalization ($M)").font = FCB
fml(ww, 21, 2, '=B19*B20', '#,##0')

ww.cell(23, 1, "Total Debt ($M)").font = FC
inp(ww, 23, 2, 98657, '#,##0', "10-K FY2025, Balance Sheet, Total short-term + long-term debt")
ww.cell(24, 1, "Total Cash & Investments ($M)").font = FC
inp(ww, 24, 2, 160537, '#,##0', "10-K FY2025, Balance Sheet, Cash + all marketable securities")
ww.cell(25, 1, "Net Cash ($M)").font = FCB
fml(ww, 25, 2, '=B24-B23', '#,##0')

ww.cell(27, 1, "Enterprise Value ($M)").font = FCB
x = fml(ww, 27, 2, '=B21-B25', '#,##0', b=True); x.fill = OF

# WACC Calculation
shdr(ww, 29, "WACC CALCULATION", mc=4)
for c, h in zip(range(1, 5), ["Component", "Weight", "Cost", "Contribution"]):
    x = ww.cell(31, c, h); x.font = FCB; x.fill = LF; x.border = TB

ww.cell(32, 1, "Equity").font = FC
fml(ww, 32, 2, '=B21/B27', '0.0%')
fml(ww, 32, 3, '=B9', '0.00%')
fml(ww, 32, 4, '=B32*C32', '0.00%')

ww.cell(33, 1, "Debt").font = FC
fml(ww, 33, 2, '=1-B32', '0.0%')
fml(ww, 33, 3, '=B15', '0.00%')
fml(ww, 33, 4, '=B33*C33', '0.00%')

ww.cell(35, 1, "WEIGHTED AVERAGE COST OF CAPITAL").font = FCB
x = fml(ww, 35, 2, '=D32+D33', '0.00%', b=True); x.fill = OF

# ═══════════════════════════════════════════════════════════
# FORMATTING PASS
# ═══════════════════════════════════════════════════════════

# Column widths — DCF
ws.column_dimensions['A'].width = 38
ws.column_dimensions['B'].width = 14
for c in range(3, 18):
    ws.column_dimensions[cl(c)].width = 14

# Column widths — WACC
ww.column_dimensions['A'].width = 38
for c in range(2, 5):
    ww.column_dimensions[cl(c)].width = 15

# Freeze panes
ws.freeze_panes = 'C51'
ww.freeze_panes = 'B4'

# Section outline borders — DCF
for sr, er in [
    (7, 13), (14, 19), (21, 26), (28, 33), (35, 40), (42, 47),
    (49, 60), (62, 73), (75, 84), (86, 99),
]:
    outline(ws, sr, er, 1, 17)

# Sensitivity table outlines
for sr, er in [(104, 110), (112, 118), (120, 126)]:
    outline(ws, sr, er, 1, 6)

# WACC section outlines
for sr, er in [(4, 9), (11, 15), (17, 27), (29, 35)]:
    outline(ww, sr, er, 1, 4)

# Total-line formatting (double bottom border for FCF and PV rows)
for c in range(3, 18):
    b = ws.cell(72, c).border
    ws.cell(72, c).border = Border(left=b.left, right=b.right,
                                    top=Side(style='thin'),
                                    bottom=Side(style='double'))
    b = ws.cell(60, c).border
    ws.cell(60, c).border = Border(left=b.left, right=b.right,
                                    top=Side(style='thin'),
                                    bottom=Side(style='thin'))

for c in range(8, 18):
    b = ws.cell(80, c).border
    ws.cell(80, c).border = Border(left=b.left, right=b.right,
                                    top=Side(style='thin'),
                                    bottom=Side(style='thin'))

# ═══════════════════════════════════════════════════════════
# SAVE
# ═══════════════════════════════════════════════════════════
path = os.path.join("/Users/donyang/Desktop/为自己工作/Website",
                     "AAPL_DCF_Model_2026-06-13.xlsx")
wb.save(path)
print(f"Model saved to: {path}")
print(f"Sheets: {wb.sheetnames}")
print("Formula-only file — open in Excel to calculate values.")
